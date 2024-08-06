# -*- coding:utf-8 -*-
from pathlib import Path

from openpyxl.reader.excel import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.workbook import Workbook

from .tools import process_content, ok_list, make_valid_name


class OriginalSetter(object):
    def __init__(self, recorder):
        self._recorder = recorder

    def cache_size(self, size):
        """设置缓存大小
        :param size: 缓存大小
        :return: None
        """
        if not isinstance(size, int) or size < 0:
            raise TypeError('cache_size值只能是int，且必须>=0')
        self._recorder._cache = size
        return self

    def path(self, path):
        """设置文件路径
        :param path: 文件路径
        :return: None
        """
        if self._recorder._path:
            self._recorder.record()

        p = Path(path)
        self._recorder._path = str(p.parent / make_valid_name(p.name))
        self._recorder._data = []
        return self

    def show_msg(self, on_off):
        """设置是否显示运行信息
        :param on_off: bool表示开关
        :return: None
        """
        self._recorder.show_msg = on_off
        return self


class BaseSetter(OriginalSetter):
    def table(self, name):
        """设置默认表名
        :param name: 表名
        :return: None
        """
        self._recorder._table = name
        return self

    def before(self, before):
        """设置在数据前面补充的列
        :param before: 列表、元组或字符串，为字符串时则补充一列
        :return: None
        """
        if before is None:
            self._recorder._before = None
        elif isinstance(before, (list, dict)):
            self._recorder._before = before
        elif isinstance(before, tuple):
            self._recorder._before = list(before)
        else:
            self._recorder._before = [before]
        return self

    def after(self, after):
        """设置在数据后面补充的列
        :param after: 列表、元组或字符串，为字符串时则补充一列
        :return: None
        """
        if after is None:
            self._recorder._after = None
        elif isinstance(after, (list, dict)):
            self._recorder._after = after
        elif isinstance(after, tuple):
            self._recorder._after = list(after)
        else:
            self._recorder._after = [after]
        return self

    def encoding(self, encoding):
        """设置编码
        :param encoding: 编码格式
        :return: None
        """
        self._recorder._encoding = encoding
        return self


class SheetLikeSetter(BaseSetter):
    def head(self, head, table=None, to_file=True):
        """设置表头。只有 csv 和 xlsx 格式支持设置表头
        :param head: 表头，列表或元组
        :param table: 表名，只xlsx格式文件有效
        :param to_file: 是否写入到文件
        :return: None
        """
        self._recorder.record()
        with self._recorder._lock:
            if not self._recorder.path:
                raise FileNotFoundError('未指定文件。')
            if not isinstance(head, (list, tuple)):
                raise TypeError('head参数只能是list或tuple格式。')

            if self._recorder.type == 'xlsx':
                table = table or self._recorder.table
                set_xlsx_head(self._recorder, head, table, to_file)

            elif self._recorder.type == 'csv':
                set_csv_head(self._recorder, head, to_file)

            else:
                raise TypeError('只能对xlsx和csv文件设置表头。')
        return self

    def delimiter(self, delimiter):
        """设置csv文件分隔符
        :param delimiter: 分隔符
        :return: None
        """
        self._recorder._delimiter = delimiter
        return self

    def quote_char(self, quote_char):
        """设置csv文件引用符
        :param quote_char: 引用符
        :return: None
        """
        self._recorder._quote_char = quote_char
        return self

    def path(self, path, file_type=None):
        """设置文件路径
        :param path: 文件路径
        :param file_type: 要设置的文件类型，为空则从文件名中获取
        :return: None
        """
        super().path(path)

        if not file_type:
            suffix = Path(path).suffix.lower()
            if suffix:
                file_type = suffix[1:]
            elif not self._recorder.type:
                file_type = 'csv'

        if file_type:
            self.file_type(file_type)

        if self._recorder._type == 'xlsx':
            self._recorder._data = {}
            self._recorder._head = {}
            self._recorder._style_data = {}
        else:
            self._recorder._data = []
            self._recorder._head = None

        return self

    def file_type(self, file_type):
        """指定文件类型，无视文件后缀名"""
        if 'any' not in self._recorder._SUPPORTS and file_type not in self._recorder._SUPPORTS:
            raise TypeError(f'只支持{"、".join(self._recorder._SUPPORTS)}格式文件。')
        self._recorder._type = file_type
        return self

    def table(self, name):
        """设置默认表名
        :param name: 表名
        :return: None
        """
        if isinstance(name, bool):
            name = None
        self._recorder._table = name
        return self

    def fit_head(self, on_off=True):
        """设置是否自动匹配表头
        :param on_off: bool表示开关
        :return: None
        """
        if self._recorder.type not in ('csv', 'xlsx'):
            raise TypeError('只有csv或xlsx格式可设置fit_head。')
        self._recorder.record()
        self._recorder._fit_head = on_off
        return self


class FillerSetter(SheetLikeSetter):
    def sign(self, value):
        """设置sign值
        :param value: 筛选条件
        :return: None
        """
        self._recorder._sign = value
        return self

    def deny_sign(self, on_off=True):
        """设置是否反向匹配sign
        :param on_off: bool表示开或关
        :return: None
        """
        self._recorder._deny_sign = on_off
        return self

    def key_cols(self, cols):
        """设置作为关键字的列，可以是多列
        :param cols: 列号或列名，或它们组成的list或tuple
        :return: None
        """
        if cols is True:
            self._recorder._key_cols = True
        elif isinstance(cols, int) and cols > 0:
            self._recorder._key_cols = [cols]
        elif isinstance(cols, str):
            self._recorder._key_cols = [int(cols)] if cols.isdigit() else [column_index_from_string(cols)]
        elif isinstance(cols, (list, tuple)):
            self._recorder._key_cols = [i if isinstance(i, int) and i > 0 else
                                        int(i) if i.isdigit() else column_index_from_string(i) for i in cols]
        else:
            raise TypeError('col值只能是int或str，且必须大于0。')
        return self

    def sign_col(self, col):
        """设置用于判断是否已填数据的列
        :param col: 列号或列名
        :return: None
        """
        if col is True or (isinstance(col, int) and col > 0):
            self._recorder._sign_col = col
        elif isinstance(col, str):
            self._recorder._sign_col = int(col) if col.isdigit() else column_index_from_string(col)
        else:
            raise TypeError('col值只能是True、int或str，且必须大于0。')
        return self

    def data_col(self, col):
        """设置用于填充数据的列
        :param col: 列号或列名
        :return: None
        """
        if isinstance(col, int) and col > 0:
            self._recorder._data_col = col
        elif isinstance(col, str):
            self._recorder._data_col = column_index_from_string(col)
        else:
            raise TypeError('col值只能是int或str，且必须大于0。')
        return self

    def begin_row(self, row):
        """设置数据开始的行
        :param row: 行号
        :return: None
        """
        if not isinstance(row, int) or row < 1:
            raise TypeError('row值只能是int，且必须大于0')
        self._recorder._begin_row = row
        return self

    def path(self, path=None, key_cols=None, begin_row=None, sign_col=None,
             data_col=None, sign=None, deny_sign=None):
        """设置文件路径
        :param path: 保存的文件路径
        :param key_cols: 作为关键字的列，可以是多列
        :param begin_row: 数据开始的行，默认表头一行
        :param sign_col: 用于判断是否已填数据的列
        :param data_col: 要填入数据的第一列
        :param sign: 按这个值判断是否已填数据
        :param deny_sign: 是否反向匹配sign，即筛选指不是sign的行
        """
        if path:
            super().path(path)
        self._recorder.set.key_cols(key_cols or self._recorder.key_cols)
        self._recorder.set.begin_row(begin_row or self._recorder.begin_row)
        self._recorder.set.sign_col(sign_col or self._recorder.sign_col)
        self._recorder.set.sign(sign or self._recorder.sign)
        self._recorder.set.data_col(data_col or self._recorder.data_col)
        self._recorder.set.deny_sign(deny_sign if deny_sign is not None else self._recorder.deny_sign)
        return self

    def link_style(self, style):
        """设置单元格的链接样式
        :param style: CellStyle对象
        :return: None
        """
        self._recorder._link_style = style
        return self


class RecorderSetter(SheetLikeSetter):
    def follow_styles(self, on_off=True):
        """设置是否跟随最后一行的style，只有xlsx格式有效
        :param on_off: True或False
        :return: None
        """
        self._recorder._follow_styles = on_off
        return self

    def col_height(self, height):
        """设置行高，只有xlsx格式有效
        :param height: 行高，传入None清空设置
        :return: None
        """
        self._recorder._col_height = height
        return self

    def styles(self, styles):
        """设置新行样式，只有xlsx格式有效，可传入多个，传入None则取消
        :param styles: 传入CellStyle对象设置整个新行，传入CellStyle对象组成的列表设置多个，传入None清空设置
        :return: None
        """
        self._recorder.record()
        self._recorder._follow_styles = False
        self._recorder._style = styles
        return self

    def path(self, path, file_type=None):
        """设置文件路径
        :param path: 文件路径
        :param file_type: 要设置的文件类型，为空则从文件名中获取
        :return: None
        """
        super().path(path=path, file_type=file_type)
        self._recorder._row_styles = None
        return self

    def fit_head(self, on_off=True, add_new=False):
        """设置是否自动匹配表头
        :param on_off: bool表示开关
        :param add_new: 数据中有表头不存在的列时是否自动添加到表头，on_off为True时才有效
        :return: None
        """
        super().fit_head(on_off)
        self._recorder._auto_new_col = add_new
        return self


class DBSetter(BaseSetter):
    def path(self, path, table=None):
        """重写父类方法
        :param path: 文件路径
        :param table: 数据表名称
        :return: None
        """
        with self._recorder._lock:
            super().path(path)
            if self._recorder._conn is not None:
                self._recorder._close_connection()
            self._recorder._connect()

            if table:
                self._recorder._table = table

            else:
                r = self._recorder.run_sql("select name from sqlite_master where type='table'")
                self._recorder._table = r[0] if r else None

            self._recorder._data = {}
            self._recorder._close_connection()
        return self


def set_csv_head(recorder, head, to_file):
    """设置csv文件的表头
    :param recorder: Recorder或Filler对象
    :param head: 表头列表或元组
    :param to_file: 是否写入文件
    :return: None
    """
    recorder._head = head
    if not to_file:
        return

    from csv import writer
    if recorder._file_exists or Path(recorder.path).exists():
        with open(recorder.path, 'r', newline='', encoding=recorder._encoding) as f:
            content = "".join(f.readlines()[1:])

        with open(recorder.path, 'w', newline='', encoding=recorder._encoding) as f:
            csv_write = writer(f, delimiter=recorder._delimiter, quotechar=recorder._quote_char)
            csv_write.writerow(ok_list(head))

        with open(recorder.path, 'a+', newline='', encoding=recorder._encoding) as f:
            f.write(f'{content}')

    else:
        Path(recorder.path).parent.mkdir(parents=True, exist_ok=True)
        with open(recorder.path, 'w', newline='', encoding=recorder._encoding) as f:
            csv_write = writer(f, delimiter=recorder._delimiter, quotechar=recorder._quote_char)
            csv_write.writerow(ok_list(head))


def set_xlsx_head(recorder, head, table, to_file):
    """设置xlsx文件的表头
    :param recorder: Recorder或Filler对象
    :param head: 表头列表或元组
    :param table: 工作表名称
    :param to_file: 是否写入文件
    :return: None
    """
    if not to_file:
        if table:
            recorder._head[table] = head
        elif recorder._file_exists or Path(recorder.path).exists():
            wb = load_workbook(recorder.path)
            ws = wb.active
            recorder._head[ws.title] = head
            wb.close()
        else:
            recorder._head['Sheet'] = head
        return

    if recorder._file_exists or Path(recorder.path).exists():
        wb = load_workbook(recorder.path)
        if table:
            ws = wb[table] if table in [i.title for i in wb.worksheets] else wb.create_sheet(title=table)
        else:
            ws = wb.active

    else:
        Path(recorder.path).parent.mkdir(parents=True, exist_ok=True)
        wb = Workbook()
        ws = wb.active
        if table:
            ws.title = table

    if len(ws[1]) > len(head):
        head = list(head)
        head.extend([None] * (len(ws[1]) - len(head)))

    for key, i in enumerate(head, 1):
        ws.cell(1, key).value = process_content(i, True)

    recorder._head[ws.title] = head
    wb.save(recorder.path)
    wb.close()
